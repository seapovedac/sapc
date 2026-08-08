#!/usr/bin/env python3
"""
generate_index.py — build the library's CSV/XLSX index.

WHAT IT DOES
    Recursively scans a PDF library folder and writes two files into it:
      - library_index.csv   plain-text index, one row per PDF
      - library_index.xlsx  same data, with the "Folder" column
                             color-coded by topic (one distinct color per
                             top-level folder, reused consistently across
                             the sheet)

    Columns: Authors | Year | Title | Document Type | Folder | Relative Path

    (There's no separate "topic" column — the top-level folder already IS
    the topic, e.g. a file in "IDP/Membrane curvature" has topic "IDP";
    a dedicated column would just duplicate what "Folder" already shows,
    so the topic is used only to pick the Folder cell's color.)

    Author/year/type/topic are inferred from each file's name and the
    top-level folder it lives in (no PDF content is read) — so it relies
    on filenames already following the library convention. Rows are
    sorted alphabetically by first author.

    Safe to re-run any time: it only reads PDF filenames and (re)writes
    the two index files — it never renames, moves, or touches any PDF.
    To fix filenames first, see format_pdf_name.py in this same folder.

USAGE
    python3 generate_index.py                  # index the folder this script lives in
    python3 generate_index.py /path/to/library  # index a specific folder
    python3 generate_index.py -h                # this help menu

REQUIREMENTS
    openpyxl (for the .xlsx output only — CSV always works):
        pip install openpyxl
"""
import argparse
import os
import re
import csv


def parse_args():
    p = argparse.ArgumentParser(
        prog="generate_index.py",
        description="Scan a PDF library and (re)generate library_index.csv / library_index.xlsx.",
        epilog=(
            "Example:\n"
            "  python3 generate_index.py /home/alejandro/Dropbox/Academia/Papers\n\n"
            "See also: format_pdf_name.py — renames/moves individual PDFs or whole\n"
            "folders of PDFs into this library's naming convention before indexing."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument(
        "library_root",
        nargs="?",
        default=os.path.dirname(os.path.abspath(__file__)),
        help="Folder to scan (default: the folder this script is in)",
    )
    return p.parse_args()


ARGS = parse_args()
LIBRARY_ROOT = ARGS.library_root

# Keep this in sync with format_pdf_name.py's TAG_CHOICES if it changes.
# [Square-bracket] tags are the canonical, machine-readable form; the
# parenthesis/no-brackets variants are legacy forms still present in
# older filenames and are matched too so this script never breaks on them.
SI_RE = re.compile(r'\[si\]|\(si\)|(^|[ _\-.])si([ _\-.]|$)|supp?(orting|lementary)', re.IGNORECASE)
EDITORIAL_RE = re.compile(r'\[editorial\]|\beditorial\b', re.IGNORECASE)
PERSPECTIVE_RE = re.compile(r'\[perspective\]|\bperspective\b', re.IGNORECASE)
NEWS_VIEWS_RE = re.compile(r'\[news\s*&?\s*views\]|news\s*&?\s*views', re.IGNORECASE)
MINI_REVIEW_RE = re.compile(r'\[mini review\]|mini[\s-]?review', re.IGNORECASE)
REVIEW_RE = re.compile(r'\[review\]|\breview\b', re.IGNORECASE)
PREPRINT_RE = re.compile(r'\[preprint\]|\bpreprint\b|biorxiv|medrxiv', re.IGNORECASE)

THESIS_RE = re.compile(r'\[(PhD|MSc) Thesis\]|\((PhD|MSc) thesis\)', re.IGNORECASE)
CHAPTER_RE = re.compile(r'\[book chapter\]|\(book chapter\)|chapter', re.IGNORECASE)
BOOK_PATTERN = re.compile(r'^(?P<authors>.+?) - (?P<title>.+?)(?: \((?P<publisher>[^)]+)\))? \((?P<year>\d{4})\)$')
ARTICLE_RE = re.compile(r'^(?P<prefix>(?:0{1,4}(?:\(\+\))?|\(\+\))[.\s]*)?(?P<author>[^\d(\[]+?)\s+(?P<year>(19|20)\d{2})')


def classify_doc_type(folder, stem):
    top = folder.split(os.sep)[0]

    if top == 'Book':
        return 'Book chapter' if CHAPTER_RE.search(stem) else 'Book'
    if top == 'Thesis' or THESIS_RE.search(stem):
        if re.search(r'PhD', stem, re.IGNORECASE):
            return 'PhD Thesis'
        if re.search(r'MSc', stem, re.IGNORECASE):
            return 'MSc Thesis'
        return 'Thesis'
    if top == 'Tutorial' or 'tutorial' in stem.lower():
        return 'Tutorial'
    if 'manual' in stem.lower():
        return 'Manual'
    if 'protocol' in stem.lower():
        return 'Protocol'
    if CHAPTER_RE.search(stem):
        return 'Book chapter'

    # document-status tags, most specific first (mirrors format_pdf_name.py's
    # guess_tag priority order)
    if SI_RE.search(stem):
        return 'Supplementary Information'
    if EDITORIAL_RE.search(stem):
        return 'Editorial'
    if PERSPECTIVE_RE.search(stem):
        return 'Perspective'
    if NEWS_VIEWS_RE.search(stem):
        return 'News & Views'
    if MINI_REVIEW_RE.search(stem):
        return 'Mini Review'
    if REVIEW_RE.search(stem):
        return 'Review'
    if PREPRINT_RE.search(stem):
        return 'Preprint'
    return 'Research article'


def parse_authors_year(stem):
    m = ARTICLE_RE.match(stem)
    if m:
        author = m.group('author').strip(' .-')
        year = m.group('year')
        return author, year
    bm = BOOK_PATTERN.match(stem)
    if bm:
        return bm.group('authors').strip(), bm.group('year')
    ym = re.search(r'(19|20)\d{2}', stem)
    year = ym.group(0) if ym else ''
    return stem.split('(')[0].strip(' .-0'), year


def topic_from_folder(folder):
    parts = folder.split(os.sep)
    return parts[0] if parts and parts[0] else 'Uncategorized'


def collect_rows():
    rows = []
    for dirpath, dirnames, filenames in os.walk(LIBRARY_ROOT):
        rel_dir = os.path.relpath(dirpath, LIBRARY_ROOT)
        if rel_dir == '.':
            rel_dir = ''
        for fname in filenames:
            if not fname.lower().endswith('.pdf'):
                continue
            stem = fname[:-4]
            doc_type = classify_doc_type(rel_dir, stem)
            authors, year = parse_authors_year(stem)
            # Not a CSV column — used only to pick the Folder cell's color.
            topic = topic_from_folder(rel_dir) if rel_dir else 'Root'
            rel_path = os.path.join(rel_dir, fname) if rel_dir else fname
            rows.append({
                'Authors': authors,
                'Year': year,
                'Title': stem,
                'Document Type': doc_type,
                'Folder': rel_dir if rel_dir else '(root)',
                'Relative Path': rel_path,
                '_topic': topic,
            })
    rows.sort(key=lambda r: (r['Authors'].lower(), r['Year']))
    return rows


def write_csv(rows, path):
    fieldnames = ['Authors', 'Year', 'Title', 'Document Type', 'Folder', 'Relative Path']
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        w.writeheader()
        w.writerows(rows)


PALETTE = [
    "FFADAD", "FFD6A5", "FDFFB6", "CAFFBF", "9BF6FF", "A0C4FF", "BDB2FF", "FFC6FF",
    "FFB4A2", "E4C1F9", "A9DEF9", "D0F4DE", "FCF6BD", "FF99C8", "B9FBC0", "8EECF5",
    "90DBF4", "F1C0E8", "CFBAF0", "A3C4F3",
]


def write_xlsx(rows, path):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Library Index"

    headers = ['Authors', 'Year', 'Title', 'Document Type', 'Folder', 'Relative Path']
    FOLDER_COL = headers.index('Folder') + 1  # 1-based for openpyxl
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    topics = sorted({r['_topic'] for r in rows})
    topic_color = {t: PALETTE[i % len(PALETTE)] for i, t in enumerate(topics)}

    for r in rows:
        ws.append([r['Authors'], r['Year'], r['Title'], r['Document Type'], r['Folder'], r['Relative Path']])
        row_idx = ws.max_row
        fill = PatternFill(start_color=topic_color[r['_topic']], end_color=topic_color[r['_topic']], fill_type="solid")
        ws.cell(row=row_idx, column=FOLDER_COL).fill = fill

    widths = [28, 6, 60, 16, 35, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"
    wb.save(path)


def main():
    rows = collect_rows()
    csv_path = os.path.join(LIBRARY_ROOT, 'library_index.csv')
    xlsx_path = os.path.join(LIBRARY_ROOT, 'library_index.xlsx')
    write_csv(rows, csv_path)
    try:
        write_xlsx(rows, xlsx_path)
    except ImportError:
        print("openpyxl not installed; skipped .xlsx generation. Run: pip install openpyxl")
    print(f"Indexed {len(rows)} PDFs.")
    print(f"CSV:  {csv_path}")
    print(f"XLSX: {xlsx_path}")


if __name__ == '__main__':
    main()
